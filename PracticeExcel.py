import openpyxl
path = "/Testing.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print("no of rows and columns in the excel are   ", rows, cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r, column=c).value, end="         ")
    print()


from openpyxl import Workbook, load_workbook
wb = load_workbook("demo.xlsx")
sh = wb['Demo1']
row_ct = sh.name
col_ct = sh.town

for i in range(1,row_ct+1):
    for j in range(1,col_ct+1):
        print(sh.cell(row=i, column=j).value)
